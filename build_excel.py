import sys, os
sys.path.insert(0, os.path.dirname(__file__))

from list import COMPANIES_PART1
from list import COMPANIES_PART2
from list import COMPANIES_PART3
from list import COMPANIES_PART4
from list import COMPANIES_PART5
from list import COMPANIES_PART6

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

# Combine and deduplicate
ALL = (COMPANIES_PART1 + COMPANIES_PART2 + COMPANIES_PART3
       + COMPANIES_PART4 + COMPANIES_PART5 + COMPANIES_PART6)

# Deduplicate by (name lowercased)
seen = set()
companies = []
for row in ALL:
    key = row[0].strip().lower()
    if key in seen:
        continue
    seen.add(key)
    companies.append(row)

print("Total unique companies: "+ str(len(companies)))

# Build workbook
wb = Workbook()

# ---------- README sheet ----------
readme = wb.active
readme.title = "README"

readme_rows = [
    ["Canadian Tech-Hiring Companies - Job Search Workbook"],
    [f"Generated: {datetime.now().strftime('%Y-%m-%d')}"],
    [f"Total companies: {len(companies)}"],
    [""],
    ["PURPOSE"],
    ["This workbook lists Canadian companies known to hire for technology roles:"],
    ["Data Scientist, Software Developer, Full Stack Developer, Frontend Developer,"],
    ["Cloud Engineer, Software Engineer, Web Developer, AI/ML Developer,"],
    ["Automation Engineer, DevOps, and related positions."],
    [""],
    ["HOW TO USE"],
    ["1. Open the 'Companies' sheet for the full list."],
    ["2. Use Excel's filter (Data > Filter) on any column to narrow by:"],
    ["   - Industry (Banking, Fintech, Gaming, HealthTech, etc.)"],
    ["   - Size (Startup, SMB, Mid, Large, Enterprise)"],
    ["   - HQ City"],
    ["3. Click the 'Careers URL' link to open the company's jobs page."],
    ["4. Click the 'LinkedIn Jobs URL' to see open LinkedIn postings."],
    ["5. Use the category sheets for quick pre-filtered views."],
    [""],
    ["NOTES"],
    ["- URLs are based on public data and may change over time."],
    ["- LinkedIn URLs use the format /company/<slug>/jobs/ - some slugs may need manual fix."],
    ["- A company being listed does NOT mean they are currently hiring - verify on their site."],
    ["- Companies with 'Remote (CA)' are global/US companies that hire Canadian remote workers."],
    [""],
    ["LEGEND - COMPANY SIZE"],
    ["Startup     : < 20 employees, early stage"],
    ["SMB / Small : 20-200 employees"],
    ["Mid         : 200-1,000 employees"],
    ["Large       : 1,000-10,000 employees"],
    ["Enterprise  : 10,000+ employees"],
    [""],
    ["TIPS FOR YOUR JOB SEARCH"],
    ["- Apply on the company's own careers site when possible (better than job boards)."],
    ["- Set LinkedIn job alerts for each target company (Follow -> Jobs -> Set alert)."],
    ["- Tailor your resume for each application - mention the exact role and company."],
    ["- Check 'Engineering blog' or 'Tech' subdomains (e.g., engineering.shopify.com)."],
    ["- For large companies, search their careers site by location: 'Canada' or specific city."],
    ["- Many Canadian employers post to Indeed, Glassdoor, and Jobbank.gc.ca too."],
]

for r in readme_rows:
    readme.append(r)

readme["A1"].font = Font(size=18, bold=True, color="1F4E79")
readme["A2"].font = Font(size=10, italic=True, color="595959")
readme["A3"].font = Font(size=11, bold=True)
for row_idx in [5, 11, 22, 29, 36]:
    c = readme.cell(row=row_idx, column=1)
    c.font = Font(bold=True, size=12, color="1F4E79")

readme.column_dimensions["A"].width = 100

# ---------- Main Companies sheet ----------
ws = wb.create_sheet("Companies")
headers = [
    "#", "Company Name", "Industry", "Size", "HQ City",
    "Website", "Careers URL", "LinkedIn Jobs URL", "Typical Tech Roles"
]
ws.append(headers)

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

for col_idx, _ in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_idx)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(vertical="center", horizontal="left")
    c.border = border

alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
link_font = Font(color="0563C1", underline="single")

for idx, (name, website, careers, slug, industry, size, city, roles) in enumerate(companies, 1):
    linkedin = f"https://www.linkedin.com/company/{slug}/jobs/" if slug else ""
    ws.append([idx, name, industry, size, city, website, careers, linkedin, roles])
    row = idx + 1
    # Links
    for col, url in [(6, website), (7, careers), (8, linkedin)]:
        cell = ws.cell(row=row, column=col)
        if url:
            cell.hyperlink = url
            cell.font = link_font
    # alternating rows
    if idx % 2 == 0:
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = alt_fill
    # Borders
    for col in range(1, len(headers) + 1):
        ws.cell(row=row, column=col).border = border

# Column widths
widths = [6, 34, 26, 12, 22, 38, 50, 58, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# ---------- Category sheets ----------
def add_category_sheet(title, filter_func):
    s = wb.create_sheet(title[:31])  # sheet name max 31 chars
    s.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = s.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.border = border
    n = 0
    for (name, website, careers, slug, industry, size, city, roles) in companies:
        if not filter_func(industry, size, city):
            continue
        n += 1
        linkedin = f"https://www.linkedin.com/company/{slug}/jobs/" if slug else ""
        s.append([n, name, industry, size, city, website, careers, linkedin, roles])
        row = n + 1
        for col, url in [(6, website), (7, careers), (8, linkedin)]:
            cell = s.cell(row=row, column=col)
            if url:
                cell.hyperlink = url
                cell.font = link_font
        if n % 2 == 0:
            for col in range(1, len(headers) + 1):
                s.cell(row=row, column=col).fill = alt_fill
        for col in range(1, len(headers) + 1):
            s.cell(row=row, column=col).border = border
    for i, w in enumerate(widths, 1):
        s.column_dimensions[get_column_letter(i)].width = w
    s.freeze_panes = "A2"
    if n > 0:
        s.auto_filter.ref = s.dimensions

add_category_sheet("Startups & SMB",
    lambda ind, size, city: size in ("Startup", "SMB", "Small"))

add_category_sheet("Mid-Size",
    lambda ind, size, city: size == "Mid")

add_category_sheet("Large & Enterprise",
    lambda ind, size, city: size in ("Large", "Enterprise"))

add_category_sheet("Fintech & Banking",
    lambda ind, size, city: any(k in ind.lower() for k in
        ["fintech", "bank", "insurance", "investment", "pension", "payments",
         "brokerage", "asset mgmt", "credit union", "insurtech", "crypto"]))

add_category_sheet("AI ML & Data",
    lambda ind, size, city: any(k in ind.lower() for k in
        ["ai", " ml", "machine", "data cloud", "analytic", "quantum"]))

add_category_sheet("Gaming & Media",
    lambda ind, size, city: any(k in ind.lower() for k in
        ["gaming", "media", "streaming", "entertainment", "animation", "film", "music"]))

add_category_sheet("HealthTech & Biotech",
    lambda ind, size, city: any(k in ind.lower() for k in
        ["health", "biotech", "pharma", "medical"]))

add_category_sheet("SaaS & Enterprise SW",
    lambda ind, size, city: any(k in ind.lower() for k in
        ["saas", "enterprise software", "crm", "erp", "edtech", "hr saas", "dev tools"]))

add_category_sheet("Toronto",
    lambda ind, size, city: "toronto" in city.lower())

add_category_sheet("Vancouver",
    lambda ind, size, city: "vancouver" in city.lower() or "burnaby" in city.lower() or "richmond, bc" in city.lower())

add_category_sheet("Montreal",
    lambda ind, size, city: "montreal" in city.lower() or "laval" in city.lower() or "quebec" in city.lower())

add_category_sheet("Calgary & Edmonton",
    lambda ind, size, city: "calgary" in city.lower() or "edmonton" in city.lower())

add_category_sheet("Ottawa & Waterloo",
    lambda ind, size, city: "ottawa" in city.lower() or "waterloo" in city.lower() or "kitchener" in city.lower())

add_category_sheet("Remote CA",
    lambda ind, size, city: "remote" in city.lower())

# Save
out = "/app/output/Canada_Tech_Companies.xlsx"
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print(f"Saved: {out}")
print(f"Sheets: {[s.title for s in wb.worksheets]}")